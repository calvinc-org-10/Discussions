import os, uuid, re as regex
import subprocess, signal
import json
from functools import partial
from django.contrib.auth.decorators import login_required
from django.db import connection, transaction
from django.http import HttpResponse
from django.shortcuts import render
from django_q.tasks import async_task
from openpyxl import load_workbook
from cMenu.models import getcParm
from cMenu.utils import ExcelWorkbook_fileext
import WICS.globals
from WICS.models import SAPPlants_org
from WICS.models import WhsePartTypes, MaterialList, tmpMaterialListUpdate
from WICS.models_async_comm import async_comm, set_async_comm_state




################################################################################
################################################################################
################################################################################

##### the suite of procs to support fnUpdateMatlListfromSAP

def proc_MatlListSAPSprsheet_00InitUMLasync_comm(reqid):
    acomm = set_async_comm_state(
        reqid,
        statecode = 'rdng-sprsht-init',
        statetext = 'Initializing ...',
        new_async=True
        )

def proc_MatlListSAPSprsheet_00CopyUMLSpreadsheet(req, reqid):
    acomm = set_async_comm_state(
        reqid,
        statecode = 'uploading-sprsht',
        statetext = 'Uploading Spreadsheet',
        )

    SAPFile = req.FILES['SAPFile']
    svdir = getcParm('SAP-FILELOC')
    fName = svdir+"tmpMatlList"+str(uuid.uuid4())+ExcelWorkbook_fileext
    with open(fName, "wb") as destination:
        for chunk in SAPFile.chunks():
            destination.write(chunk)

    return fName

def proc_MatlListSAPSprsheet_01ReadSpreadsheet(reqid, fName):
acomm = set_async_comm_state(
    reqid,
    statecode = 'rdng-sprsht',
    statetext = 'Reading Spreadsheet',
    )

# tmpMaterialListUpdate does multiple duty: it will store the MM60 spreadsheet, and it will identify 
# what Material as found in WICS, what needs to be added, and what WICS material is no longer in SAP
# tmpMaterialListUpdate is also used at the end to report results back to the user
tmpMaterialListUpdate.objects.all().delete()

SAP_SSName_TableName_map = {
        # dictionary mapping 'name of column in spreadsheet':'name of column in tmpMaterialListUpdate table'
        'Material': 'Material', 
        'Plant': 'Plant',
        # etc
        }

wb = load_workbook(filename=fName, read_only=True)
ws = wb.active
SAPcol = {'Plant':None,'Material': None}
SAPcolmnNames = ws[1]
# after this loop, SAPcol['name of col in tmpMaterialListUpdate table'] = spreadsheet col# it's in 
for col in SAPcolmnNames:
    if col.value in SAP_SSName_TableName_map:
        SAPcol[SAP_SSName_TableName_map[col.value]] = col.column - 1
if (SAPcol['Material'] == None or SAPcol['Plant'] == None):
    set_async_comm_state(
        reqid,
        statecode = 'fatalerr',
        statetext = 'SAP Spreadsheet has bad header row. Plant and/or Material is missing.  See Calvin to fix this.',
        result = 'FAIL - bad spreadsheet',
        )

    wb.close()
    os.remove(fName)
    return

numrows = ws.max_row
nRows = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    nRows += 1
    if nRows % 100 == 0:
        set_async_comm_state(
            reqid,
            statecode = 'rdng-sprsht',
            statetext = f'Reading Spreadsheet ... record {nRows} of {numrows}',
            )

    # examine the row, adding it to tmpMaterialListUpdate, along with an error message if needed
    if row[SAPcol['Material']]==None: MatNum = ''
    else: MatNum = row[SAPcol['Material']]
    ## refuse to work with special chars embedded in the MatNum
    if regex.match(".*[\n\t\xA0].*",MatNum):
        tmpMaterialListUpdate(
            recStatus = 'err-MatlNum',
            errmsg = f'error: {MatNum!a} is an unusable part number. It contains invalid characters and cannot be added to WICS',
            Material = row[SAPcol['Material']], 
            # MaterialLink = MaterialLink,
            Description = row[SAPcol['Description']], 
            Plant = row[SAPcol['Plant']],
            # etc
            ).save()
        continue
    elif len(str(MatNum)):
        _org = SAPPlants_org.objects.filter(SAPPlant=row[SAPcol['Plant']])[0].org
        tmpMaterialListUpdate(
            org = _org,
            Material = row[SAPcol['Material']], 
            # MaterialLink = MaterialLink,
            Description = row[SAPcol['Description']], 
            Plant = row[SAPcol['Plant']],
            # etc
            ).save()
    # endif invalid Material 
# endfor
wb.close()
os.remove(fName)
def done_MatlListSAPSprsheet_01ReadSpreadsheet(t):
reqid = t.args[0]
statecode = async_comm.objects.get(pk=reqid).statecode
if statecode != 'fatalerr':
    set_async_comm_state(
        reqid,
        statecode = 'done-rdng-sprsht',
        statetext = f'Finished Reading Spreadsheet',
        )
    # this call is here because 02 CANNOT start before 01 is finished
    task02 = async_task(proc_MatlListSAPSprsheet_02_identifyexistingMaterial, reqid, hook=done_MatlListSAPSprsheet_02_identifyexistingMaterial)
    # the hook of 02(identify) will trigger 03(Remove) and 04(Add) and 99(End)
#endif stateocde != 'fatalerr'

def proc_MatlListSAPSprsheet_02_identifyexistingMaterial(reqid):
set_async_comm_state(
    reqid,
    statecode = 'get-matl-link',
    statetext = f'Finding SAP MM60 Materials already in WICS Material List',
    )
UpdMaterialLinkSQL = 'UPDATE WICS_tmpmateriallistupdate, (select id, org_id, Material from WICS_materiallist) as MasterMaterials'
UpdMaterialLinkSQL += ' set WICS_tmpmateriallistupdate.MaterialLink_id = MasterMaterials.id, '
UpdMaterialLinkSQL += "     WICS_tmpmateriallistupdate.recStatus = 'FOUND' "
UpdMaterialLinkSQL += ' where WICS_tmpmateriallistupdate.org_id = MasterMaterials.org_id '
UpdMaterialLinkSQL += '   and WICS_tmpmateriallistupdate.Material = MasterMaterials.Material '
# tmpMaterialListUpdate.objects.all().update(MaterialLink=Subquery(MaterialList.objects.filter(org=OuterRef('org'), Material=OuterRef('Material'))[0]))
with connection.cursor() as cursor:
    cursor.execute(UpdMaterialLinkSQL)
set_async_comm_state(
    reqid,
    statecode = 'get-matl-link-done',
    statetext = f'Finished linking SAP MM60 list to existing WICS Materials',
    )
def done_MatlListSAPSprsheet_02_identifyexistingMaterial(t):
reqid = t.args[0]

task03 = async_task(proc_MatlListSAPSprsheet_03_Remove, reqid,)
task04 = async_task(proc_MatlListSAPSprsheet_04_Add, reqid,)

def proc_MatlListSAPSprsheet_03_Remove(reqid):
set_async_comm_state(
    reqid,
    statecode = 'id-del-matl',
    statetext = f'Identifying WICS Materials no longer in SAP MM60 Materials',
    )
MustKeepMatlsCond = []
MustKeepMatlsCond.append(('.SEL.','id NOT IN (SELECT DISTINCT MaterialLink_id AS Material_id FROM WICS_tmpmateriallistupdate WHERE MaterialLink_id IS NOT NULL)'))
MustKeepMatlsCond.append(('.DEL.','(org_id, Material) IN (SELECT DISTINCT org_id, Material FROM WICS_tmpmateriallistupdate WHERE recStatus like "DEL%")'))
MustKeepMatlsCond.append(('.SEL.','id NOT IN (SELECT DISTINCT Material_id FROM WICS_actualcounts)'))
MustKeepMatlsCond.append(('.SEL.','id NOT IN (SELECT DISTINCT Material_id FROM WICS_countschedule)'))
MustKeepMatlsCond.append(('.SEL.','id NOT IN (SELECT DISTINCT Material_id FROM WICS_sap_sohrecs)'))
MustKeepMatlsSelCond = ''
MustKeepMatlsDelCond = ''
for sqlsttyp, phr in MustKeepMatlsCond:
    if 'SEL' in sqlsttyp:
        if MustKeepMatlsSelCond: MustKeepMatlsSelCond += ' AND '
        MustKeepMatlsSelCond += f'({phr})'
    if 'DEL' in sqlsttyp:
        if MustKeepMatlsDelCond: MustKeepMatlsDelCond += ' AND '
        MustKeepMatlsDelCond += f'({phr})'

DeleteMatlsSelectSQL = "INSERT INTO WICS_tmpmateriallistupdate (recStatus, MaterialLink_id, org_id, Material, Description, Plant )"
DeleteMatlsSelectSQL += " SELECT  concat('DEL ',FORMAT(id,0)), NULL, org_id, Material, Description, Plant "
DeleteMatlsSelectSQL += " FROM WICS_materiallist"
DeleteMatlsSelectSQL += f" WHERE ({MustKeepMatlsSelCond})"
with connection.cursor() as cursor:
    cursor.execute(DeleteMatlsSelectSQL)

set_async_comm_state(
    reqid,
    statecode = 'del-matl-2',
    statetext = f'Removing WICS Materials no longer in SAP MM60 Materials',
    )
# do the Removals
DeleteMatlsDoitSQL = "DELETE FROM WICS_materiallist"
DeleteMatlsDoitSQL += f" WHERE ({MustKeepMatlsDelCond})"
with connection.cursor() as cursor:
    cursor.execute(DeleteMatlsDoitSQL)
    transaction.on_commit(partial(done_MatlListSAPSprsheet_03_Remove,reqid))
def done_MatlListSAPSprsheet_03_Remove(reqid):
key = f'MatlX{reqid}'
statecodeVal = ".03."
if async_comm.objects.filter(pk=key).exists():
    MatlXval = async_comm.objects.get(pk=key).statecode + statecodeVal
else:
    MatlXval = statecodeVal
set_async_comm_state(
    key, 
    statecode = MatlXval,
    statetext = '',
    new_async = True
    )
set_async_comm_state(
    reqid,
    statecode = 'del-matl-done',
    statetext = f'Finished Removing WICS Materials no longer in SAP MM60 Materials',
    )

def proc_MatlListSAPSprsheet_04_Add(reqid):
set_async_comm_state(
    reqid,
    statecode = 'id-add-matl',
    statetext = f'Identifying SAP MM60 Materials new to WICS',
    )

# first pass, for presentation in results - orgname rather than org
MarkAddMatlsSelectSQL = "UPDATE WICS_tmpmateriallistupdate"
MarkAddMatlsSelectSQL += " SET recStatus = 'ADD'"
MarkAddMatlsSelectSQL += " WHERE (MaterialLink_id IS NULL) AND (recStatus is NULL)"
with connection.cursor() as cursor:
    cursor.execute(MarkAddMatlsSelectSQL)

set_async_comm_state(
    reqid,
    statecode = 'add-matl',
    statetext = f'Adding SAP MM60 Materials new to WICS',
    )
UnknownTypeID = 18  # internally used code
# do the adds
# one day django will implement insert ... select.  Until then ...
AddMatlsSelectSQL = "SELECT"
AddMatlsSelectSQL += f" org_id, Material, Description, Plant, {UnknownTypeID} AS PartType_id"
AddMatlsSelectSQL += " FROM WICS_tmpmateriallistupdate"
AddMatlsSelectSQL += " WHERE (MaterialLink_id IS NULL) AND (recStatus = 'ADD') "

AddMatlsDoitSQL = "INSERT INTO WICS_materiallist"
AddMatlsDoitSQL += " (org_id, Material, Description, Plant, PartType_id"
AddMatlsDoitSQL += ")"
AddMatlsDoitSQL += f' {AddMatlsSelectSQL}'
with connection.cursor() as cursor:
    cursor.execute(AddMatlsDoitSQL)

set_async_comm_state(
    reqid,
    statecode = 'add-matl-get-recid',
    statetext = f'Getting Record ids of SAP MM60 Materials new to WICS',
    )
UpdMaterialLinkSQL = 'UPDATE WICS_tmpmateriallistupdate, (select id, org_id, Material from WICS_materiallist) as MasterMaterials'
UpdMaterialLinkSQL += ' set WICS_tmpmateriallistupdate.MaterialLink_id = MasterMaterials.id '
UpdMaterialLinkSQL += ' where WICS_tmpmateriallistupdate.org_id = MasterMaterials.org_id '
UpdMaterialLinkSQL += '   and WICS_tmpmateriallistupdate.Material = MasterMaterials.Material '
UpdMaterialLinkSQL += "   and (MaterialLink_id IS NULL) AND (recStatus = 'ADD')"
with connection.cursor() as cursor:
    cursor.execute(UpdMaterialLinkSQL)
    transaction.on_commit(partial(done_MatlListSAPSprsheet_04_Add,reqid))
def done_MatlListSAPSprsheet_04_Add(reqid):
key = f'MatlX{reqid}'
statecodeVal = ".04."
if async_comm.objects.filter(pk=key).exists():
    MatlXval = async_comm.objects.get(pk=key).statecode + statecodeVal
else:
    MatlXval = statecodeVal
set_async_comm_state(
    key, 
    statecode = MatlXval,
    statetext = '',
    new_async = True
    )
set_async_comm_state(
    reqid,
    statecode = 'add-matl-done',
    statetext = f'Finished Adding SAP MM60 Materials new to WICS',
    )

def proc_MatlListSAPSprsheet_99_FinalProc(reqid):
set_async_comm_state(
    reqid,
    statecode = 'done',
    statetext = 'Finished Processing Spreadsheet',
    )

def proc_MatlListSAPSprsheet_99_Cleanup(reqid):
# also kill reqid, acomm, qcluster process
async_comm.objects.filter(pk=reqid).delete()
os.kill(int(reqid), signal.SIGTERM)
os.kill(int(reqid), signal.SIGKILL)

# delete the temporary table
tmpMaterialListUpdate.objects.all().delete()

@login_required
def fnUpdateMatlListfromSAP(req):

client_phase = req.POST['phase'] if 'phase' in req.POST else None
reqid = req.COOKIES['reqid'] if 'reqid' in req.COOKIES else None

if req.method == 'POST':
    # check if the mandatory commits have been done and change the status code if so
    if reqid is not None:
        mandatory_commit_key = f'MatlX{reqid}'
        mandatory_commit_list = ['03', '04']
        if async_comm.objects.filter(pk=mandatory_commit_key).exists():
            mandatory_commits_recorded = async_comm.objects.get(pk=mandatory_commit_key).statecode
            if all((c in mandatory_commits_recorded) for c in mandatory_commit_list):
                proc_MatlListSAPSprsheet_99_FinalProc(reqid)
                async_comm.objects.filter(pk=mandatory_commit_key).delete()

    if client_phase=='init-upl':
        retinfo = HttpResponse()

        # start django_q broker
        reqid = subprocess.Popen(
            "python manage.py qcluster"
        ).pid
        # reqid = uuid.uuid4()
        retinfo.set_cookie('reqid',str(reqid))
        proc_MatlListSAPSprsheet_00InitUMLasync_comm(reqid)

        fName = proc_MatlListSAPSprsheet_00CopyUMLSpreadsheet(req, reqid)
        task01 = async_task(proc_MatlListSAPSprsheet_01ReadSpreadsheet, reqid, fName, hook=done_MatlListSAPSprsheet_01ReadSpreadsheet)

        acomm_fake = {
            'statecode': 'starting',
            'statetext': 'SAP MM60 Update Starting',
            }
        retinfo.write(json.dumps(acomm_fake))
        return retinfo
    elif client_phase=='waiting':
        retinfo = HttpResponse()

        acomm = async_comm.objects.values().get(pk=reqid)    # something's very wrong if this doesn't exist
        stcode = acomm['statecode']
        if stcode == 'fatalerr':
            pass
        retinfo.write(json.dumps(acomm))
        return retinfo
    elif client_phase=='wantresults':
        ImpErrList = tmpMaterialListUpdate.objects.filter(recStatus__startswith='err')
        AddedMatlsList = tmpMaterialListUpdate.objects.filter(recStatus='ADD')
        RemvdMatlsList = tmpMaterialListUpdate.objects.filter(recStatus__startswith='DEL')
        cntext = {
            'ImpErrList':ImpErrList,
            'AddedMatls':AddedMatlsList,
            'RemvdMatls':RemvdMatlsList,
            }
        templt = 'frmUpdateMatlListfromSAP_done.html'
        return render(req, templt, cntext)
    elif client_phase=='resultspresented':
        proc_MatlListSAPSprsheet_99_Cleanup(reqid)
        retinfo = HttpResponse()
        retinfo.delete_cookie('reqid')

        return retinfo
    else:
        return
    #endif client_phase
else:
    cntext = {
        }
    templt = 'frmUpdateMatlListfromSAP_phase0.html'
#endif req.method = 'POST'

return render(req, templt, cntext)
